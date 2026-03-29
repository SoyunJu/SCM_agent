"""
Excel 업로드 엔드포인트 단위 테스트
- 파일 검증, 파싱 성공/실패, DB 업서트 연동 검증
"""
import io
import pytest
from unittest.mock import patch, MagicMock
from fastapi import FastAPI
from fastapi.testclient import TestClient


# ── Fixture ──────────────────────────────────────────────────────────────────

def _make_app():
    from app.api.sheets_router import router
    from app.api.auth_router import get_current_user, require_admin, TokenData
    from app.db.connection import get_db

    app = FastAPI()
    app.include_router(router)
    mock_user = TokenData(username="admin", role="admin")
    app.dependency_overrides[require_admin]    = lambda: mock_user
    app.dependency_overrides[get_current_user] = lambda: mock_user
    app.dependency_overrides[get_db]           = lambda: MagicMock()
    return app


@pytest.fixture
def client():
    return TestClient(_make_app())


def _make_xlsx_bytes() -> bytes:
    """상품마스터 테스트용 xlsx"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "상품마스터"
    ws.append(["상품코드", "상품명", "카테고리", "안전재고기준"])
    ws.append(["P001", "테스트상품A", "카테고리1", 10])
    ws.append(["P002", "테스트상품B", "카테고리2", 5])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_sales_xlsx_bytes() -> bytes:
    """일별판매 테스트용 xlsx"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "일별판매"
    ws.append(["날짜", "상품코드", "판매수량", "매출액"])
    ws.append(["2026-03-01", "P001", 5, 50000])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 정상 업로드 ──────────────────────────────────────────────────────────────

def test_upload_master_success(client):
    """상품마스터 업로드 성공 — inserted/updated 반환"""
    xlsx_bytes = _make_xlsx_bytes()

    with patch("app.api.sheets_router.upsert_master_from_excel", return_value=None), \
            patch("app.api.sheets_router.SyncService") as mock_sync:

        mock_sync.sync_master.return_value = {"inserted": 2, "updated": 0, "skipped": 0}

        resp = client.post(
            "/scm/sheets/upload-excel",
            files={"file": ("test.xlsx", xlsx_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"sheet_type": "master"},
        )

    assert resp.status_code == 200
    body = resp.json()
    assert body["status"] == "success"
    assert body["sheet_type"] == "master"
    assert body["total"] == 2


def test_upload_sales_success(client):
    """일별판매 업로드 성공"""
    xlsx_bytes = _make_sales_xlsx_bytes()

    with patch("app.api.sheets_router.write_sales", return_value=None), \
            patch("app.api.sheets_router.SyncService") as mock_sync:

        mock_sync.sync_sales.return_value = {"inserted": 1, "updated": 0, "skipped": 0}

        resp = client.post(
            "/scm/sheets/upload-excel",
            files={"file": ("sales.xlsx", xlsx_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"sheet_type": "sales"},
        )

    assert resp.status_code == 200
    assert resp.json()["status"] == "success"
    assert resp.json()["total"] == 1


def test_upload_stock_success(client):
    """재고현황 업로드 성공"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["상품코드", "현재재고", "입고예정일", "입고예정수량"])
    ws.append(["P001", 50, "2026-04-01", 100])
    buf = io.BytesIO()
    wb.save(buf)

    with patch("app.api.sheets_router.upsert_stock_from_excel", return_value=None), \
            patch("app.api.sheets_router.SyncService") as mock_sync:

        mock_sync.sync_stock.return_value = {"inserted": 1, "updated": 0, "skipped": 0}

        resp = client.post(
            "/scm/sheets/upload-excel",
            files={"file": ("stock.xlsx", buf.getvalue(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"sheet_type": "stock"},
        )

    assert resp.status_code == 200
    assert resp.json()["status"] == "success"


# ── 유효성 검증 오류 ─────────────────────────────────────────────────────────

def test_upload_invalid_extension(client):
    """xlsx 이외 확장자 → 400"""
    resp = client.post(
        "/scm/sheets/upload-excel",
        files={"file": ("test.csv", b"col1,col2", "text/csv")},
        data={"sheet_type": "master"},
    )
    assert resp.status_code == 400
    assert "xlsx" in resp.json()["detail"]


def test_upload_invalid_sheet_type(client):
    """유효하지 않은 sheet_type → 400"""
    xlsx_bytes = _make_xlsx_bytes()
    resp = client.post(
        "/scm/sheets/upload-excel",
        files={"file": ("test.xlsx", xlsx_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"sheet_type": "invalid"},
    )
    assert resp.status_code == 400
    assert "sheet_type" in resp.json()["detail"]


def test_upload_missing_required_columns(client):
    """필수 컬럼 누락 xlsx → 422"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "상품마스터"
    ws.append(["잘못된컬럼"])
    ws.append(["P001"])
    buf = io.BytesIO()
    wb.save(buf)

    resp = client.post(
        "/scm/sheets/upload-excel",
        files={"file": ("bad.xlsx", buf.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"sheet_type": "master"},
    )
    assert resp.status_code == 422
    assert "필수 컬럼 누락" in resp.json()["detail"]